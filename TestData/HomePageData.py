import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname": "Krithi", "lastname": "Shetty", "Gender": "Female"},
                          {"firstname": "Shailendar", "lastname": "Reddy", "Gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\shailendar reddy\\PycharmProjects\\PythonDemo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    val = sheet.cell(row=i, column=j).value
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        print(Dict)
        print("You Indian Guy fuck off")
        return[Dict]

