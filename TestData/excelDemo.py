import openpyxl

book = openpyxl.load_workbook("C:\\Users\\shailendar reddy\\PycharmProjects\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=2, column=3)
print(cell.value)

sheet.cell(row=2, column=3).value = "reddy"
print(sheet.cell(row=2, column=3).value)

print(sheet.max_column)
print(sheet.max_row)
print(sheet["D2"].value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase1":
        for j in range(2, sheet.max_column+1):
            val = sheet.cell(row=i, column=j).value
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)












